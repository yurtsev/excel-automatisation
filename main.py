import os
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import openpyxl as op
from datetime import datetime

filepath = ''


def data_match():
    global date_time_only_date, date_value_date
    date_time_value = sheetBAM['A' + str(non_empty_cells_ZES + 1)].value
    print(date_time_value)
    date_value = sheet["A5"].value

    # Преобразуем значения в дату (если они не уже в формате datetime)
    if not isinstance(date_time_value, datetime):
        date_time_value = datetime.strptime(date_time_value, '%Y-%m-%d %H:%M:%S')

    if not isinstance(date_value, datetime):
        date_value = datetime.strptime(date_value, '%Y-%m-%d')

    # Извлекаем только дату из значения с датой и временем
    date_time_only_date = date_time_value.date()  # cтатистика
    date_value_date = date_value.date()
    # Сравниваем даты
    if date_time_only_date == date_value_date:
        return True
    else:
        return False


def btn_rechoice():
    global filepath, sheet
    filepath = filedialog.askopenfilename(title="Выбор файла", initialdir='E:\Потребление',
                                          filetypes=[("Excel files", "*.xls;*.xlsx")])
    filename = os.path.basename(filepath)
    readdate2.configure(text='"' + filename + '"', bg="black", foreground='white')
    readdate2.place(x=10, y=155)
    readfile = op.load_workbook(filepath)
    sheet = readfile["Набор"]
    date_value = sheet["A5"].value
    if data_match() == True:
        readdate3.configure(text=str(date_value_date), bg="black", foreground="lightgreen",
                            font=("Arial", 12))
        readdate3.place(x=10, y=175)
        writedate3.configure(text=str(date_time_only_date), bg="black", foreground="lightgreen",
                             font=("Arial", 12))
        writedate3.place(x=10, y=85)
    else:
        readdate3.configure(text=str(date_value_date), bg="black", foreground="red",
                            font=("Arial", 12))
        readdate3.place(x=10, y=175)
        writedate3.configure(text=str(date_time_only_date), bg="black", foreground="red",
                             font=("Arial", 12))
        writedate3.place(x=10, y=85)


def btn_choice():
    global filepath, filename
    filepath = filedialog.askopenfilename(title="Выбор файла", initialdir='D:\Потребление',
                                          filetypes=[("Excel files", "*.xls;*.xlsx")])
    filename = os.path.basename(filepath)
    if filepath == '':
        errorlabel.configure(text="файл не выбран", bg="black", foreground='white')
        errorlabel.pack()
    else:
        errorlabel.configure(text=filename, justify=RIGHT, foreground='white')
        errorlabel.pack()


def btn_save():
    global combobam, combozes, errorlabel2, windowzb
    print(filepath)
    if filepath == '':
        errorlabel.configure(text='нужно выбрать файл!', foreground='red')
        errorlabel.pack()
    else:
        root.destroy()
        windowzb = Tk()
        windowzb.title("выбор колонок")
        screen_widthzb = windowzb.winfo_screenwidth()
        screen_heightzb = windowzb.winfo_screenheight()

        x = (screen_widthzb - 300) // 2
        y = (screen_heightzb - 160) // 3
        windowzb.geometry("300x160+" + str(x) + '+' + str(y))
        windowzb.resizable(False, False)
        zbframe = Frame(windowzb, bg="black")
        zbframe.place(width=300, height=160)
        zblabel = Label(zbframe, text="Укажите колонку, в которой \n находятся данные по ЗЭС и БАМ", bg="black",
                        foreground='white', font=("Arial", 12))
        zblabel.pack()
        zblabel2 = Label(zbframe, text="ЗЭС", bg="black",
                         foreground='white', font=("Arial", 12))
        zblabel2.place(x=20, y=70)
        zblabel3 = Label(zbframe, text="БАМ", bg="black",
                         foreground='white', font=("Arial", 12))
        zblabel3.place(x=170, y=70)
        errorlabel2 = Label(zbframe, text="выберите обе колонки!", bg="black",
                            foreground='black', font=("Arial", 12))
        errorlabel2.pack()
        style = ttk.Style()
        style.theme_use('default')

        style.configure('TCombobox',
                        fieldbackground='black',
                        background='white')
        columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
        combozes = ttk.Combobox(zbframe, values=columns, width=3, font=("Arial", 12), foreground='white')
        combozes.place(x=72, y=71)
        combobam = ttk.Combobox(zbframe, values=columns, width=3, font=("Arial", 12), foreground='white')
        combobam.place(x=224, y=71)
        btn_confim = Button(zbframe, text="продолжить", bg="white", width=15, command=general)
        btn_confim.place(y=115, x=93)


def general():
    global BAMcolumn, ZEScolumn, sheetZES, sheetBAM, sheet, writefile, non_empty_cells_ZES, alarmwindow
    alarmwindow = None
    BAMcolumn = combobam.get()
    ZEScolumn = combozes.get()
    print(BAMcolumn, '---', ZEScolumn)
    if BAMcolumn != '' and ZEScolumn != '':

        readfile = op.load_workbook(filepath)
        sheet = readfile["Набор"]

        writefile = op.load_workbook("D:\Потребление\Статистика по территории 2021-2024.xlsx")
        sheetZES = writefile["база_ЗЭС"]
        sheetBAM = writefile["база_БАМ"]

        column_letter = 'B'  # Например, столбец A

        # Считаем количество непустых строк в столбце
        non_empty_cells_ZES = 0
        for cell in sheetZES[column_letter]:
            if cell.value is not None:
                non_empty_cells_ZES += 1  # работает

        if data_match() == True:  # data_value - дата выбранного файла
            windowzb.destroy()
            exel()
        else:

            alarmgo()
    else:
        print(BAMcolumn, '---', ZEScolumn)
        errorlabel2.configure(text="выберите обе колонки!", bg="black",
                              foreground='red', font=("Arial", 12))
        errorlabel2.pack()

def alarmgo():
    global writedate2, writedate3, readdate2, readdate3, alarmwindow
    alarmwindow = Tk()
    windowzb.destroy()
    alarmwindow.title('!!!')
    alarmwindow['bg'] = 'black'
    screen_widthrt = alarmwindow.winfo_screenwidth()
    screen_heightrt = alarmwindow.winfo_screenheight()
    x = (screen_widthrt - 550) // 2
    y = (screen_heightrt - 225) // 3
    alarmwindow.geometry('550x225+' + str(x) + '+' + str(y))
    alarmwindow.resizable(False, False)
    alarmframe = Frame(alarmwindow, bg='black')
    alarmframe.place(width=550, height=225)
    date = Label(alarmframe, text="даты не совпадают!", bg="black", foreground='white', font=("Arial", 16))
    date.place(x=10, y=5)
    writedate = Label(alarmframe, text="дата файла", bg="black", foreground='white', font=("Arial", 12))
    writedate.place(x=10, y=45)
    writedate2 = Label(alarmframe, text='"Статистика по территории 2021-2024.xlsx"', bg="black",
                       foreground="white", font=("Arial", 12))
    writedate2.place(x=10, y=65)
    writedate3 = Label(alarmframe, text=str(date_time_only_date), bg="black", foreground="red",
                       font=("Arial", 12))
    writedate3.place(x=10, y=85)

    readdate = Label(alarmframe, text="дата файла", bg="black", foreground='white', font=("Arial", 12))
    readdate.place(x=10, y=135)
    readdate2 = Label(alarmframe, text='"' + filename + '"', bg="black", foreground="white", font=("Arial", 12))
    readdate2.place(x=10, y=155)
    readdate3 = Label(alarmframe, text=str(date_value_date), bg="black", foreground="red",
                      font=("Arial", 12))
    readdate3.place(x=10, y=175)

    asklabel = Label(alarmframe, text="продолжить?", bg="black", foreground='white', font=("Arial", 14))
    asklabel.place(x=376, y=40)

    btn_go = Button(alarmframe, text="продолжить", bg="white", height=2, width=15, command=exel)
    btn_go.place(x=380, y=85)

    btn_no = Button(alarmframe, text="изменить файл", bg="white", height=2, width=15, command=btn_rechoice)
    btn_no.place(x=380, y=140)

def exel():
    global alarmwindow
    if alarmwindow is not None and alarmwindow.winfo_exists():
        alarmwindow.destroy()

    column_letter = 'B'  # Например, столбец A

    # Считаем количество непустых строк в столбце
    non_empty_cells_ZES = 0
    for cell in sheetZES[column_letter]:
        if cell.value is not None:
            non_empty_cells_ZES += 1  # работает
    print(non_empty_cells_ZES)

    if data_match() == True:  # data_value - дата выбранного файла
        row_count = (sheet.max_row - 4) // 24 * 24 + 4
        day = (sheet.max_row - 4) // 24
        print(row_count, day)
        for i in range(1, day + 1):
            for j in range((i - 1) * 24 + 5, i * 24 + 5):
                print(sheet[ZEScolumn + str(j)].value, '-=-', non_empty_cells_ZES + i, '-=-', (j - 24 * (i - 1) - 3))
                sheetZES.cell(row=non_empty_cells_ZES + i, column=(j-24*(i-1)-3)).value = sheet[ZEScolumn + str(j)].value
                sheetBAM.cell(row=non_empty_cells_ZES + i, column=(j - 24 * (i - 1) - 3)).value = sheet[BAMcolumn + str(j)].value
    writefile.save("D:\Потребление\Статистика по территории 2021-2024.xlsx")



root = Tk()
root.title('ЗЭС и БАМ')
root['bg'] = 'black'
screen_widthrt = root.winfo_screenwidth()
screen_heightrt = root.winfo_screenheight()
x = (screen_widthrt - 200) // 2
y = (screen_heightrt - 225) // 3
root.geometry('200x225+' + str(x) + '+' + str(y))
root.resizable(False, False)
frame = Frame(root, bg='black')
frame.place(width=150, height=200, x=25, y=25)
btn_ch = Button(frame, text="выбрать файл", bg="white", command=btn_choice)
btn_ch.pack(ipady=5, fill=X)

errorlabel = Label(frame, text="файл не выбран", bg="black", foreground='white')
errorlabel.pack(ipady=20, fill=X)

btn_rech = Button(frame, text="изменить файл", bg="white", command=btn_choice)
btn_rech.pack(ipady=5, fill=X)
btn_save = Button(frame, text="сохранить выбор", bg="white", command=btn_save)
btn_save.pack(ipady=5, fill=X, pady=10)

root.mainloop()
